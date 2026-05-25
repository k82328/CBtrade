#!/usr/bin/env python3
"""
CB 三大法人日買賣超爬蟲 v7
- 若當天無資料，自動往前找最近有資料的交易日（最多回溯5天）
- 同時輸出 cb_inst_YYYYMMDD.xlsx 和 cb_inst_latest.json
"""

import argparse, time, json, requests, pandas as pd
from datetime import datetime, timedelta

API = "https://www.tpex.org.tw/www/zh-tw/bond/newCb3itrade"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    "Referer": "https://www.tpex.org.tw/zh-tw/bond/info/statistics-cb/inst-trading/day.html",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://www.tpex.org.tw",
}

def to_datetime(s): return datetime.strptime(s, "%Y/%m/%d")

def fetch_one_day(dt):
    label = dt.strftime("%Y/%m/%d")
    payload = {"type": "Daily", "date": label, "id": "", "response": "json"}
    try:
        r = requests.post(API, data=payload, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        if data.get("stat") != "ok":
            return pd.DataFrame()
        tables = data.get("tables", [])
        if not tables:
            return pd.DataFrame()
        rows = tables[0].get("data", [])
        if not rows:
            return pd.DataFrame()
        return parse_rows(rows, label)
    except Exception as e:
        print(f"    ⚠️  {e}")
        return pd.DataFrame()

def fetch_with_fallback(dt, max_lookback=5):
    """若當天無資料，自動往前找最近有資料的交易日"""
    checked = 0
    cur = dt
    while checked < max_lookback:
        # 跳過週末
        if cur.weekday() >= 5:
            cur -= timedelta(days=1)
            continue
        label = cur.strftime("%Y/%m/%d")
        print(f"  嘗試 {label}...", end=" ", flush=True)
        df = fetch_one_day(cur)
        if not df.empty:
            if cur < dt:
                print(f"✅ {len(df)} 筆（回溯至 {label}）")
            else:
                print(f"✅ {len(df)} 筆")
            return df
        else:
            print("無資料，往前一日")
            cur -= timedelta(days=1)
            checked += 1
        time.sleep(0.8)
    print(f"  ❌ 往前 {max_lookback} 個交易日均無資料")
    return pd.DataFrame()

def c(v):
    try: return int(str(v).replace(",","").replace("--","0").strip() or "0")
    except: return 0

def parse_rows(rows, label):
    records = []
    for row in rows:
        records.append({
            "日期": label, "CB代號": str(row[0]).strip(), "CB名稱": str(row[1]).strip(),
            "外資_買張": c(row[2]), "外資_賣張": c(row[3]), "外資_買賣超": c(row[4]),
            "投信_買張": c(row[5]), "投信_賣張": c(row[6]), "投信_買賣超": c(row[7]),
            "自營_買張": c(row[8]), "自營_賣張": c(row[9]), "自營_買賣超": c(row[10]),
            "三大法人合計買賣超": c(row[11]) if len(row)>11 else c(row[4])+c(row[7])+c(row[10]),
        })
    return pd.DataFrame(records)

def recent_trading_days(n):
    days, dt = [], datetime.today()
    while len(days) < n:
        if dt.weekday() < 5: days.append(dt)
        dt -= timedelta(days=1)
    return list(reversed(days))

def save_excel(df, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="三大法人CB買賣超")
        ws = writer.sheets["三大法人CB買賣超"]
        for col in ws.columns:
            w = max((len(str(cell.value)) for cell in col if cell.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w+2, 28)
        ws.freeze_panes = "A2"
        from openpyxl.styles import PatternFill
        green = PatternFill("solid", fgColor="C6EFCE")
        red   = PatternFill("solid", fgColor="FFC7CE")
        if "三大法人合計買賣超" in df.columns:
            ci = df.columns.get_loc("三大法人合計買賣超") + 1
            for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                for cell in row:
                    try:
                        v = int(cell.value or 0)
                        if v > 0: cell.fill = green
                        elif v < 0: cell.fill = red
                    except: pass
    print(f"✅ Excel：{path}")

def save_json(df, path):
    records = df.to_dict(orient="records")
    out = {
        "updated": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "date": df["日期"].iloc[0] if len(df) > 0 else "",
        "data": records,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON：{path}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date")
    parser.add_argument("--days", type=int, default=1)
    parser.add_argument("--start")
    parser.add_argument("--end")
    args = parser.parse_args()

    if args.date:
        targets = [to_datetime(args.date)]
    elif args.start:
        s = to_datetime(args.start)
        e = to_datetime(args.end) if args.end else datetime.today()
        targets = [s+timedelta(i) for i in range((e-s).days+1) if (s+timedelta(i)).weekday()<5]
    else:
        targets = recent_trading_days(args.days)

    print(f"查詢 {len(targets)} 個交易日...")
    frames = []
    for dt in targets:
        df = fetch_with_fallback(dt)
        if not df.empty:
            frames.append(df)
        time.sleep(1.0)

    if not frames:
        print("\n❌ 無資料")
        return

    result = pd.concat(frames, ignore_index=True)
    result["買超方向"] = result["三大法人合計買賣超"].apply(
        lambda v: "↑買超" if v>0 else ("↓賣超" if v<0 else "持平")
    )
    result = result.sort_values(["日期","三大法人合計買賣超"], ascending=[True,False])

    dates = result["日期"].unique()
    tag = dates[0].replace("/","") if len(dates)==1 else f"{dates[0].replace('/','')}_to_{dates[-1].replace('/','')}"

    save_excel(result, f"cb_inst_{tag}.xlsx")

    # 最新一天輸出 JSON 供網頁使用
    latest = result[result["日期"]==dates[-1]]
    save_json(latest, "cb_inst_latest.json")

    cols = ["CB代號","CB名稱","外資_買賣超","投信_買賣超","自營_買賣超","三大法人合計買賣超"]
    print(f"\n📊 {dates[-1]} 買超前5：")
    print(latest.nlargest(5,"三大法人合計買賣超")[cols].to_string(index=False))
    print(f"\n📊 {dates[-1]} 賣超前5：")
    print(latest.nsmallest(5,"三大法人合計買賣超")[cols].to_string(index=False))

if __name__ == "__main__":
    main()
